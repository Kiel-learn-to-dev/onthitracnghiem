"""Import the existing Excel question bank into SQLite."""

from __future__ import annotations

import hashlib
import json
import sqlite3
from pathlib import Path

import openpyxl

from scripts.storage import record_canonical_question, record_source, record_source_question


def _checksum(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _question_hash(question: str, choices: list[str]) -> str:
    payload = json.dumps([question, choices], ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def import_excel(db_path: Path, workbook_path: Path) -> None:
    """Import all Excel rows in one transaction, preserving their source form."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    questions_sheet, answers_sheet = workbook.worksheets
    answers = {
        int(row[0]): (str(row[1]).strip().upper(), str(row[2]).strip())
        for row in answers_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }

    connection = sqlite3.connect(db_path)
    try:
        connection.execute("PRAGMA foreign_keys = ON")
        with connection:
            source_id = record_source(
                connection, workbook_path.name, "xlsx", _checksum(workbook_path)
            )
            for row in questions_sheet.iter_rows(min_row=2, values_only=True):
                ordinal, topic, difficulty, question, *choices = row
                ordinal = int(ordinal)
                normalized_choices = [str(choice).strip() for choice in choices]
                answer, explanation = answers[ordinal]
                choices_json = json.dumps(normalized_choices, ensure_ascii=False)
                canonical_id = record_canonical_question(
                    connection,
                    str(question).strip(),
                    choices_json,
                    answer,
                    explanation,
                    str(topic).strip(),
                    str(difficulty).strip(),
                    _question_hash(str(question).strip(), normalized_choices),
                )
                record_source_question(
                    connection,
                    source_id,
                    ordinal,
                    str(question).strip(),
                    choices_json,
                    canonical_id,
                )
    finally:
        connection.close()
        workbook.close()

